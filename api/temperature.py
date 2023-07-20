from api.database import get_cursor
from datetime import datetime

from typing import List, Dict
from models.TemperaturesList import TemperaturesList
from models.TemperatureModel import TemperatureModel
from models.TemperaturePostModel import TemperaturePostModel
from models.Success import Success

from openpyxl import Workbook
from openpyxl.styles import NamedStyle
from random import randint
from hashlib import md5

def post_temperature(request: TemperaturePostModel):
    cursor = get_cursor()
    current_ts = datetime.now()

    insert_temperature_sql = '''
        INSERT INTO temperatures (pseudo_table_id, temperature, thermometer_id, ts, lamp_enabled)
        VALUES (?, ?, ?, ?, ?)
    '''
    for temperature in request.temperatures:
        args = (request.pseudo_table_id, temperature.value, temperature.thermometer_id, current_ts, request.lamp_enabled)
        cursor.execute(insert_temperature_sql, args)
    
    cursor.connection.commit()
    cursor.connection.close()

    return Success(success=True)

def remove_pseudo_table(pseudo_table_id: str):
    cursor = get_cursor()
    remove_pseudo_table_sql = '''
    DELETE FROM temperatures WHERE pseudo_table_id = ?
    '''
    cursor.execute(remove_pseudo_table_sql, (pseudo_table_id,))
    cursor.connection.commit()
    cursor.connection.close()

    return Success(success=True) 

def get_temperatures(pseudo_table_id: str):
    cursor = get_cursor()
    select_times_indexes_sql = '''
        SELECT pseudo_table_id, ts FROM temperatures WHERE pseudo_table_id = ? GROUP BY ts
    '''
    cursor.execute(select_times_indexes_sql, (pseudo_table_id,))
    times_indexes: List[str] = map(lambda row: row["ts"], cursor.fetchall())

    result = TemperaturesList()
    result.temps = {}

    for time in times_indexes:
        print(time)
        select_temps_sql = '''
            SELECT temperature, thermometer_id FROM temperatures WHERE pseudo_table_id = ? AND ts = ?
        '''
        args = (pseudo_table_id, time)
        key = time.replace(" ", "T")

        cursor.execute(select_temps_sql, args)
        result.temps[key] = list()

        temps: List[dict] = cursor.fetchall()

        for temp in temps:
            result.temps[key].append(TemperatureModel(thermometer_id=temp["thermometer_id"], value=temp["temperature"]))
        
    cursor.connection.close()
    return result

def get_pseudo_tables():
    cursor = get_cursor()
    get_pseudo_tables_sql = '''
        SELECT pseudo_table_id FROM temperatures WHERE 1 GROUP BY pseudo_table_id
    '''
    cursor.execute(get_pseudo_tables_sql)
    l = list()
    for r in cursor.fetchall():
        l.append(r["pseudo_table_id"])
        
    cursor.connection.close()
    return l

def construct_excel(pseudo_table_id: str):
    cursor = get_cursor()

    wb = Workbook()

    # grab the active worksheet
    ws = wb.active

    ws['A1'] = 'Time / Lamp'

    get_timestamps_sql = '''SELECT ts, lamp_enabled FROM temperatures WHERE pseudo_table_id = ? ORDER BY ts ASC'''
    cursor.execute(get_timestamps_sql, (pseudo_table_id,))
    tss = cursor.fetchall()
    index_ts = 2

    knownts = []

    for ts in tss:
        if ts["ts"] in knownts:
            continue
        knownts.append(ts["ts"])
        cellkey__ = 'A' + str(index_ts)
        cellkey__lamp = 'B' + str(index_ts)
        cell = ws[cellkey__]
        cell.value = datetime.fromisoformat(ts["ts"]).strftime("%H:%M:%S")
        
        cell = ws[cellkey__lamp]
        cell.value = ts["lamp_enabled"]

        index_ts += 1

    get_thermos_sql = '''
        SELECT thermometer_id FROM temperatures WHERE pseudo_table_id = ? GROUP BY thermometer_id
    '''
    cursor.execute(get_thermos_sql, (pseudo_table_id,))
    cols = cursor.fetchall()


    # adding thermometer id columns
    thermo_column = ['C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N']
    thermo_index = 0
    for col in cols:
        ws.column_dimensions[thermo_column[thermo_index]].width = 20.43
        cellkey = thermo_column[thermo_index] + '1'
        ws[cellkey] = "Thermometer #%(id)s" % {
            'id': col["thermometer_id"]
        }

        get_thermo_vals_sql = '''SELECT temperature FROM temperatures WHERE pseudo_table_id = ? AND thermometer_id = ? ORDER BY ts ASC'''
        cursor.execute(get_thermo_vals_sql, (pseudo_table_id, col["thermometer_id"]))
        vals = cursor.fetchall()
        val_row_index = 2
        for row in vals:
            cellkey_ = thermo_column[thermo_index] + str(val_row_index)
            ws[cellkey_] = row["temperature"]
            val_row_index += 1


        thermo_index += 1
    
    cursor.connection.close()

    random_path = md5(
        str(
            randint(111111111, 9999999999)
        ).encode()
    ).hexdigest() + '.xlsx'

    wb.save(random_path)
    return random_path